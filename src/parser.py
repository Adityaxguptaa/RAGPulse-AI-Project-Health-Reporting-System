"""Excel workbook parser for project plan files."""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.models import Project, Risk, Task

logger = logging.getLogger(__name__)

# ── Helpers ──────────────────────────────────────────────────────────────────

_DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y",
    "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
]


def _to_date(value: Any) -> Optional[date]:
    """Coerce a cell value to a Python date, or return None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date
        try:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(value)
            return dt.date() if isinstance(dt, datetime) else dt
        except Exception:
            return None
    text = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(str(value).replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _cell_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _col_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    """Return {normalised_header: col_index} for a header row."""
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        if raw:
            key = re.sub(r"[\s\-_/]+", "_", str(raw).lower().strip()).rstrip("_")
            mapping[key] = col
    return mapping


def _find_header_row(ws: Worksheet, keywords: list[str], max_search: int = 10) -> int:
    """Return the first row (1-indexed) that looks like a header."""
    for row in range(1, min(max_search + 1, ws.max_row + 1)):
        row_values = [
            str(ws.cell(row, c).value or "").lower()
            for c in range(1, min(ws.max_column + 1, 30))
        ]
        matches = sum(1 for kw in keywords if any(kw in v for v in row_values))
        if matches >= 2:
            return row
    return 1


# ── Sheet-level extractors ───────────────────────────────────────────────────

def _extract_summary(ws: Worksheet) -> dict[str, Any]:
    """Pull key/value pairs from a summary-style sheet."""
    data: dict[str, Any] = {}
    for row in ws.iter_rows(values_only=True):
        key_cell = next((c for c in row if c is not None), None)
        if key_cell is None:
            continue
        key = re.sub(r"[:\s]+", "_", str(key_cell).lower().strip()).strip("_")
        if not key:
            continue
        values = [c for c in row if c is not None]
        value = values[1] if len(values) > 1 else None
        if key and value is not None:
            data[key] = value
    return data


def _extract_comments(ws: Worksheet) -> list[str]:
    """Collect all non-empty strings from a comments sheet."""
    comments: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            text = _cell_text(cell)
            if text and len(text) > 3:
                comments.append(text)
    return comments


def _extract_tasks(ws: Worksheet) -> list[Task]:
    """Parse a task/schedule sheet into Task objects."""
    header_row = _find_header_row(
        ws,
        keywords=["task", "name", "start", "finish", "status", "complete", "milestone"],
    )
    cols = _col_map(ws, header_row)

    logger.debug("Task sheet '%s' cols: %s", ws.title, list(cols.keys()))

    # Flexible column aliases
    def _get(col_aliases: list[str]) -> Optional[int]:
        for alias in col_aliases:
            norm = re.sub(r"[\s\-_/]+", "_", alias.lower())
            for key, idx in cols.items():
                if norm in key or key in norm:
                    return idx
        return None

    c_name     = _get(["task", "name", "activity", "wbs", "description", "work_package"])
    c_status   = _get(["status", "state"])
    c_bl_start = _get(["baseline_start", "planned_start", "bl_start", "base_start"])
    c_bl_fin   = _get(["baseline_finish", "planned_finish", "bl_finish", "base_finish", "target"])
    c_act_start = _get(["actual_start", "act_start", "started"])
    c_act_fin  = _get(["actual_finish", "act_finish", "completed", "end"])
    c_pct      = _get(["percent", "complete", "pct", "%_complete", "progress"])
    c_var      = _get(["variance", "var_days", "schedule_variance", "delay"])
    c_float    = _get(["float", "slack", "total_float"])
    c_critical = _get(["critical", "crit"])
    c_milestone = _get(["milestone"])
    c_notes    = _get(["notes", "remarks", "comment"])
    c_deps     = _get(["predecessor", "depend", "dep"])

    tasks: list[Task] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        def v(col: Optional[int]) -> Any:
            return ws.cell(row_idx, col).value if col else None

        name = _cell_text(v(c_name))
        if not name:
            continue

        # Detect milestone from name or dedicated column
        is_milestone = False
        if c_milestone:
            is_milestone = bool(v(c_milestone))
        if not is_milestone:
            is_milestone = any(kw in name.lower() for kw in ["milestone", "gate", "go-live", "go live", "launch"])

        # Detect critical path
        is_critical = False
        if c_critical:
            raw_crit = str(v(c_critical) or "").lower()
            is_critical = raw_crit in {"yes", "true", "1", "critical", "y"}

        pct = _to_float(v(c_pct))
        # If percent > 1 and <= 100, treat as percentage; if <= 1, treat as fraction
        if pct is not None and pct <= 1.0:
            pct = round(pct * 100, 1)

        deps: list[str] = []
        dep_raw = _cell_text(v(c_deps))
        if dep_raw:
            deps = [d.strip() for d in re.split(r"[,;|]+", dep_raw) if d.strip()]

        task = Task(
            name=name,
            status=_cell_text(v(c_status)),
            baseline_start=_to_date(v(c_bl_start)),
            baseline_finish=_to_date(v(c_bl_fin)),
            actual_start=_to_date(v(c_act_start)),
            actual_finish=_to_date(v(c_act_fin)),
            planned_finish=_to_date(v(c_bl_fin)),
            percent_complete=pct,
            variance_days=_to_float(v(c_var)),
            float_days=_to_float(v(c_float)),
            is_milestone=is_milestone,
            is_critical=is_critical,
            dependencies=deps,
            notes=_cell_text(v(c_notes)),
        )
        tasks.append(task)

    return tasks


def _extract_risks(ws: Worksheet) -> list[Risk]:
    """Parse a risk register sheet."""
    header_row = _find_header_row(
        ws,
        keywords=["risk", "description", "severity", "impact", "mitigation"],
    )
    cols = _col_map(ws, header_row)

    def _get(aliases: list[str]) -> Optional[int]:
        for alias in aliases:
            norm = re.sub(r"[\s\-_/]+", "_", alias.lower())
            for key, idx in cols.items():
                if norm in key or key in norm:
                    return idx
        return None

    c_desc  = _get(["risk", "description", "issue", "name"])
    c_sev   = _get(["severity", "priority", "rating"])
    c_prob  = _get(["probability", "likelihood", "prob"])
    c_impact = _get(["impact", "consequence"])
    c_mit   = _get(["mitigation", "action", "response"])
    c_owner = _get(["owner", "responsible"])
    c_status = _get(["status", "state"])

    risks: list[Risk] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        def v(col: Optional[int]) -> Any:
            return ws.cell(row_idx, col).value if col else None

        desc = _cell_text(v(c_desc))
        if not desc:
            continue
        risks.append(Risk(
            description=desc,
            severity=_cell_text(v(c_sev)),
            probability=_cell_text(v(c_prob)),
            impact=_cell_text(v(c_impact)),
            mitigation=_cell_text(v(c_mit)),
            owner=_cell_text(v(c_owner)),
            status=_cell_text(v(c_status)),
        ))
    return risks


# ── Sheet classifier ─────────────────────────────────────────────────────────

_SUMMARY_KEYWORDS  = {"summary", "overview", "project_info", "project_details", "header"}
_COMMENT_KEYWORDS  = {"comment", "note", "issue", "log", "minutes"}
_RISK_KEYWORDS     = {"risk", "risk_register", "issue_register", "risks"}
_TASK_KEYWORDS     = {"task", "plan", "schedule", "gantt", "wbs", "activity",
                      "deliverable", "workplan", "timeline", "milestone"}


def _classify_sheet(name: str) -> str:
    """Return 'summary' | 'comments' | 'risks' | 'tasks' | 'unknown'."""
    lower = name.lower().replace(" ", "_").replace("-", "_")
    if any(kw in lower for kw in _SUMMARY_KEYWORDS):
        return "summary"
    if any(kw in lower for kw in _COMMENT_KEYWORDS):
        return "comments"
    if any(kw in lower for kw in _RISK_KEYWORDS):
        return "risks"
    if any(kw in lower for kw in _TASK_KEYWORDS):
        return "tasks"
    return "tasks"  # Default: try to parse as task sheet


# ── Legacy .xls parser ───────────────────────────────────────────────────────

def _parse_xls_workbook(path: Path) -> "Project":
    """Parse a legacy .xls file by converting it to openpyxl via pandas."""
    import pandas as pd
    import openpyxl as _xl

    # Use pandas + xlrd to read all sheets, then write to an in-memory xlsx
    import io
    xls = pd.ExcelFile(str(path), engine="xlrd")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet, header=None)
                df.to_excel(writer, sheet_name=sheet[:31], index=False, header=False)
            except Exception as e:
                logger.warning("Could not convert sheet '%s' from .xls: %s", sheet, e)

    buf.seek(0)
    wb = _xl.load_workbook(buf, data_only=True)

    project = Project(name=path.stem, source_file=str(path))
    all_tasks: list[Task] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = _classify_sheet(sheet_name)
        try:
            if sheet_type == "summary":
                meta = _extract_summary(ws)
                project.raw_metadata.update(meta)
            elif sheet_type == "comments":
                project.comments.extend(_extract_comments(ws))
            elif sheet_type == "risks":
                project.risks.extend(_extract_risks(ws))
            elif sheet_type == "tasks":
                all_tasks.extend(_extract_tasks(ws))
        except Exception as exc:
            logger.warning("Error parsing .xls sheet '%s': %s", sheet_name, exc)

    project.tasks = [t for t in all_tasks if not t.is_milestone]
    project.milestones = [t for t in all_tasks if t.is_milestone]
    project.name = path.stem.replace("_", " ").replace("-", " ").title()
    return project


# ── Public API ────────────────────────────────────────────────────────────────

def parse_workbook(file_path: str | Path) -> Project:
    """Parse an Excel project plan workbook into a Project object."""
    path = Path(file_path)
    logger.info("Parsing workbook: %s", path.name)

    try:
        # openpyxl handles .xlsx/.xlsm; fall back to xlrd for legacy .xls
        if path.suffix.lower() == ".xls":
            try:
                import xlrd
                import pandas as pd
                return _parse_xls_workbook(path)
            except ImportError:
                raise ValueError("xlrd is required to open legacy .xls files. Install it with: pip install xlrd")
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

    project = Project(name=path.stem, source_file=str(path))
    all_tasks: list[Task] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = _classify_sheet(sheet_name)
        logger.debug("Sheet '%s' classified as '%s'", sheet_name, sheet_type)

        try:
            if sheet_type == "summary":
                meta = _extract_summary(ws)
                project.raw_metadata.update(meta)
                # Attempt to pull well-known fields
                for key in ("project_name", "project", "name", "title"):
                    if key in meta:
                        project.name = str(meta[key])
                        break
                for key in ("pm", "project_manager", "manager", "owner"):
                    if key in meta:
                        project.pm = str(meta[key])
                        break
                for key in ("start", "start_date", "project_start"):
                    if key in meta:
                        project.start = _to_date(meta[key])
                        break
                for key in ("finish", "end", "end_date", "project_end", "project_finish"):
                    if key in meta:
                        project.finish = _to_date(meta[key])
                        break

            elif sheet_type == "comments":
                project.comments.extend(_extract_comments(ws))

            elif sheet_type == "risks":
                project.risks.extend(_extract_risks(ws))

            elif sheet_type == "tasks":
                tasks = _extract_tasks(ws)
                all_tasks.extend(tasks)

        except Exception as exc:
            logger.warning("Error parsing sheet '%s': %s", sheet_name, exc)

    # Split tasks vs milestones
    project.tasks = [t for t in all_tasks if not t.is_milestone]
    project.milestones = [t for t in all_tasks if t.is_milestone]

    # Fall back: if no project name from summary, use filename
    if not project.name or project.name == path.stem:
        project.name = path.stem.replace("_", " ").replace("-", " ").title()

    logger.info(
        "Parsed '%s': %d tasks, %d milestones, %d risks, %d comments",
        project.name, len(project.tasks), len(project.milestones),
        len(project.risks), len(project.comments),
    )
    return project


def load_all_workbooks(directory: str | Path) -> list[Project]:
    """Load all .xlsx / .xls workbooks from a directory."""
    directory = Path(directory)
    projects: list[Project] = []

    patterns = ["*.xlsx", "*.xls"]
    files: list[Path] = []
    for pattern in patterns:
        files.extend(sorted(directory.glob(pattern)))

    if not files:
        logger.warning("No Excel files found in '%s'", directory)
        return projects

    for f in files:
        try:
            project = parse_workbook(f)
            projects.append(project)
        except Exception as exc:
            logger.error("Failed to parse '%s': %s", f.name, exc)

    return projects
